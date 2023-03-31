import requests 
import json
import html

import os
import sys

# Get the current working directory
cwd = os.path.dirname(os.path.realpath(__file__))
import time

from bs4 import BeautifulSoup

import nltk
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize
from nltk import data
import marqo


import openai

# Redirect standard input to null device
sys.stdin = open(os.devnull, 'r')

from flask import Flask, request, jsonify
import numpy as np
from sklearn.cluster import KMeans
from gensim.models import KeyedVectors
from tqdm import tqdm

app = Flask(__name__)

# Load the pre-trained word2vec model from gensim-data
print("Loading model...")
# Check for the text file containing the model path
if os.path.exists("model_path.txt"):
    # Read the model path from the text file
    with open("model_path.txt", "r") as f:
        model_path = f.read()
else:
    # Download the model from the gensim-data repository, save the path to a text file, and read the path from the text file
    import gensim.downloader as api
    model_path = api.load("glove-wiki-gigaword-300", return_path=True)
    with open("model_path.txt", "w") as f:
        f.write(model_path)
# Load the model
model = KeyedVectors.load_word2vec_format(model_path, binary=False)
print("Model loaded.")

# sample prompt
#prompt = "Why does this java program keep crashing with the errorcode: unirestexception?"

# Define a method to aquire a prompt and keywords from the user
def get_keywords(prompt):

    # tokenize the prompt
    tokens = word_tokenize(prompt)

    # remove stop words
    stop_words = set(stopwords.words('english'))
    filtered_tokens = [word for word in tokens if word.lower() not in stop_words]

    # apply POS tagging
    pos_tags = nltk.pos_tag(filtered_tokens)

    # apply NER
    ner_tags = nltk.ne_chunk(pos_tags)

    # extract keywords based on POS and NER
    keywords = []
    for word, pos in pos_tags:
        print(word, pos, ner_tags)
        if pos.startswith('NN') or pos.startswith('JJ') or pos.startswith('VB') or 'PERSON' in str(ner_tags) or 'ORGANIZATION' in str(ner_tags):
            keywords.append(word)


    # Check if any keywords were contiguously extracted
    if len(keywords) > 0:
        # For all keywords in the list, join them into a single string
        keywords = ' '.join(keywords)

        # Print the keywords
        print("Keyword list : "+keywords)
    
    # If no keywords were extracted, ask the user for input again
    else:
        print("No keywords were extracted.")

    # Return the keywords
    return keywords

def cluster_words(words, clusters, model):
    print("words : "+words)
    # If the words are a string, split them into a list of words
    if isinstance(words, str):
        words = words.split(" ")

    # Create a list of embedding vectors for each keyword
    vectors = []
    valid_words = []
    for word in tqdm(words, desc='Processing words'):
        if word in model.key_to_index:
            vectors.append(model[word])
            valid_words.append(word)

    # Check if the vectors list is empty
    if not vectors:
        print("No valid words found in the model.")
        return {}

    all_groups = []
    for i in range(3):
        # Use KMeans to cluster the embedding vectors
        kmeans = KMeans(n_clusters=clusters+i, random_state=i).fit(vectors)

        # Group the keywords based on their assigned cluster
        cluster_labels = kmeans.labels_
        groups = {}
        for label, word in zip(cluster_labels, valid_words):
            if label not in groups:
                groups[label] = []
            groups[label].append(word)
        
        all_groups.append(groups)

    # Remove duplicate answers
    unique_groups = []
    for group in all_groups:
        if group not in unique_groups:
            unique_groups.append(group)

    # Join all the similar groups into a string
    result = []
    for group in unique_groups:
        for label in group:
            result.append(" ".join(group[label]))

    # Return the result
    print("result : ")
    print(result)
    return result

# Define a method to parse the html response and return plain text
def html_to_plain_text(html):
    # Replace all <code> tags with markdown backticks
    html = html.replace('<code>', '```').replace('</code>', '```')

    soup = BeautifulSoup(html, 'html.parser')
    plain_text = []
    for element in soup.contents:
        if element.name == 'code':
            plain_text.append('`{}`'.format(element.text))
        elif element.name:
            plain_text.append(element.get_text())
    return ' '.join(plain_text).strip()

# Define a method to get the top 5 questions from stackoverflow
def get_questions(keywords):
    # Search for questions with query "java float error"
    search_endpoint = "https://api.stackexchange.com/2.3/search/advanced"
    search_parameters = {
        "site": "stackoverflow",
        "q": keywords,
        "sort": "relevance",
        "order": "desc",
        #Include the body field in the response
        "filter": "withbody",
        # Limit to 15 results
        "pagesize": 15,
    }

    search_response = requests.get(search_endpoint, params=search_parameters)
    search_data = search_response.json()

    print(search_data)
    # If the items list is empty, cluster the keywords and try again
    if search_data.get('items') and search_data["items"].__len__() == 0:
        for i in range(0, keywords.count(" ")):

            # Get the groups from the response JSON object
            keywords = cluster_words(keywords, i+1, model)

            # For each string in keywords, search for questions with the string
            for keyword in keywords:
                print("Searching : "+keyword)
                search_parameters["q"] = keyword
                search_response = requests.get(search_endpoint, params=search_parameters)
                # Change search_response from json to a dictionary
                search_response = search_response.json()

                # If the items list exists and is greater than 0, print all the titles and extend the search_data.
                # Else, check the quota remaining. If it is 0, break and make stop sending api calls
                if search_response.get('items') and search_response["items"].__len__() > 0:
                    for item in search_response["items"]:
                        print(item["title"])
                    # Append the results to the items list
                    search_data["items"].extend(search_response["items"])
                    print("Response total : "+str(len(search_data["items"])))
                    print("Quota remaining : "+str(search_response["quota_remaining"]))
                else:
                    if search_response["error_name"] == "throttle_violation":
                        break
            # If the items list is not empty, break out of the loop. Else, continue to the next iteration with a small delay
            if search_data["items"].__len__() > 0:
                break
            else:
                time.sleep(1)
    elif search_data["error_name"] == "throttle_violation":
        print("Errors : "+str(search_response))
            
    # If the quota remaining is above 0, attempt to get answers for the questions
    if not search_data.get('error_name'):
        # Get a list of question ids from the search results
        question_ids = []
        for question in search_data["items"]:
            question_ids.append(question["question_id"])
        id_amount = len(question_ids)
        print("ID total length : "+str(id_amount))

        # Join the question ids with semicolon
        question_ids_str = ";".join(str(id) for id in question_ids)

        # Get answers for those question ids
        answer_endpoint = f"https://api.stackexchange.com/2.3/questions/{question_ids_str}/answers"
        answer_parameters = {
            "site": "stackoverflow",
            "pagesize" : 100,
            "sort": "votes",
            "order": "desc",
            "filter": "withbody",
            "page": 1
        }

        has_more = True
        answer_data = None

        # Checks for additional answers, moves on to the next page if so
        while has_more:
            answer_response = requests.get(answer_endpoint, params=answer_parameters)
            if not answer_data:
                answer_data = answer_response.json()
            else:
                answer_data["items"].extend(answer_response.json()["items"])
            has_more = answer_response.json()["has_more"]
            # ...
            answer_parameters["page"] += 1

        # Get only highest score answer for each question id
        # This must check per question ID, as multiple answers may be are for each question
        seen_questions = set()
        highest_score_answers = []
        for answer in answer_data["items"]:
            if answer["question_id"] not in seen_questions:
                highest_score_answers.append(answer)
                seen_questions.add(answer["question_id"])

        print("ID total : "+str(len(answer_data["items"])))
        print("Seen ID's : "+str(len(seen_questions)))
        return search_data, highest_score_answers
        # If the quota remaining is 0, return None for the search data and highest score answers
    else:
        print("Errors : "+str(search_response))
        return None, None

import pandas as pd
import openpyxl
# Define a method to create an excel file containing the questions and answers
def create_excel(qa_list, prompt):
    # Create an excel spreadsheet containing a column of answers and questions from the previous api call that go with it

    # Create a pandas dataframe from the list of dictionaries
    df = pd.DataFrame(qa_list)
    
    # Get the current working directory and create a folder called "Answers" if it doesn't exist
    if not os.path.exists(cwd+"/Answers"):
        os.mkdir(cwd+"/Answers")

    # Save the dataframe as an excel file, with the first column as the index and the name of the file as the prompt, in the current directories folder "Answers"

    #writer = pd.ExcelWriter(cwd+f"/Answers/{prompt}.xlsx")
    #for sheetname, dfi in df.items():  # loop through `dict` of dataframes
    #    dfi.to_excel(writer, sheet_name=sheetname)  # send df to writer
    #    worksheet = writer.sheets[sheetname]  # pull worksheet object
    #    for idx, col in enumerate(dfi):  # loop through all columns
    #        series = dfi[col]
    #        max_len = max((
    #            series.astype(str).map(len).max(),  # len of largest item
    #            len(str(series.name))  # len of column name/header
    #            )) + 1  # adding a little extra space
    #        worksheet.set_column(idx, idx, max_len)  # set column width
    #writer.save()
    
    # Modify the prompt to remove any invalid characters, replace spaces with underscores, and only take up to 31 characters
    name = prompt.replace("?", "").replace("/", "").replace("\\", "").replace("*", "").replace("[", "").replace("]", "").replace(":", "").replace(";", "").replace(",", "").replace(".", "").replace(" ", "_")[:31]
    with pd.ExcelWriter(cwd+f"/Answers/{name}.xlsx") as writer:
        df.to_excel(writer, index=False)
    
    # Load the workbook and select the worksheet
    wb = openpyxl.load_workbook(cwd+f"/Answers/{name}.xlsx")
    ws = wb.active

    # Get the maximum length of the content in each column
    max_length_a = max([len(str(cell.value)) for cell in ws["A"]])
    max_length_b = max([len(str(cell.value)) for cell in ws["B"]])
    max_length_c = max([len(str(cell.value)) for cell in ws["C"]])

    # Set the column width to the maximum length of the content + a small buffer
    ws.column_dimensions["A"].width = max_length_a + 2
    ws.column_dimensions["B"].width = (max_length_b*.75) + 2
    ws.column_dimensions["C"].width = max_length_c + 2

    # Save the changes to the workbook
    wb.save(cwd+f"/Answers/{name}.xlsx")
    
    # Save all the data in a master json file, creating it if it doesnt exist
    if not os.path.exists(cwd+"/Answers/master.json"):
        with open(cwd+"/Answers/master.json", "w") as f:
            json.dump({"items": []}, f)
    with open(cwd+"/Answers/master.json", "r") as f:
        master_data = json.load(f)
    master_data["items"].extend(qa_list)
    with open(cwd+"/Answers/master.json", "w") as f:
        json.dump(master_data, f)

# Define a method that uses vectors to see which response is the most relevant to the user's input
from sklearn.feature_extraction.text import TfidfVectorizer


def get_relevant_response(prompt, search_data, highest_score_answers):
    # Load the marqo model from docker (url is the url of the docker container)
    mq = marqo.Client(url='http://192.168.1.48:8882')

    # If highest_score_answers is not none, send the answers to the semantic search model to add them to index. Else, just send the prompt
    if highest_score_answers:
        # Create a list of dictionaries containing question and answer data
        qa_list = []
        for answer in highest_score_answers:
            qa_dict = {"Question": "", "Answer": "", "Link": ""}
            for question in search_data["items"]:
                if answer["question_id"] == question["question_id"]:
                    qa_dict["Question"] = html.unescape(question["title"])
                    qa_dict["Link"] = question["link"]
                    break
            # Convert the html response to plain text making sure its not empty
            if answer["body"]:
                qa_dict["Answer"] = html_to_plain_text(answer["body"])
            else:
                qa_dict["Answer"] = "No answer"
            qa_list.append(qa_dict)

        # Filter out any duplicate data sets
        seen_questions = set()
        unrepetitive_list = []
        for answer_list in qa_list:
            if answer_list["Question"] not in seen_questions:
                unrepetitive_list.append(answer_list)
                seen_questions.add(answer_list["Question"])
            else:
                print("Duplicate question found: "+answer_list["Question"])

        print("Adding "+str(unrepetitive_list)+" to the index")
        mq.index("best_response").add_documents(unrepetitive_list)
    else:
        print("No answers from stack")

    results = mq.index("best_response").search(
        q=prompt, searchable_attributes=["Question"], limit=30
    )

    # Filter out any duplicate data sets
    seen_questions = set()
    unrepetitive_results = []
    for new_answer_list in results['hits']:
        if new_answer_list["Question"] not in seen_questions and html.unescape(new_answer_list["Question"]) == new_answer_list["Question"]:
            # ^^^ If the Question returns different after being html unescaped, then it is a duplicate
            # Append the results["Question"], results["Answer"], results["Link"], and results["_score"] (as ["Score"]) to the list
            unrepetitive_results.append({"Question": new_answer_list["Question"], "Answer": new_answer_list["Answer"], "Link": new_answer_list["Link"] , "Score": new_answer_list["_score"]})
            #Add the question to the seen_questions set
            seen_questions.add(new_answer_list["Question"])
        else:
            print("Duplicate question found: "+new_answer_list["Question"])
            # Delete the item from the list using its _id value
            mq.index("best_response").delete_documents(ids=[new_answer_list["_id"]])

    # For all of the answers, increase their score based on their length. Above 1000 characters, the score increases slower.
    for answer in unrepetitive_results:
        length = len(answer["Answer"])
        if length > 1000:
            thousands = length//1000 
            answer["Score"] *= (1.02) + ((length-1000)/(10000*thousands))
        elif length > 50:
            answer["Score"] *= 1+((length/50)*.001)
    # Sort the qa_list by the score
    unrepetitive_results.sort(key=lambda x: x.get("Score", 0), reverse=True)

    # Print the final results
    print("Final results: ")
    print(unrepetitive_results)
    # return the qa_list
    return unrepetitive_results

# Define a method to organize the top two questions and answers into a json format to send to the front end
def get_best_questions(dictlist):
    # Get the first two lists
    bestlist = []
    runneruplist = []
    for i, qa_dict in enumerate(dictlist):
        if bestlist == []:
            bestlist = qa_dict
        elif runneruplist == []:
            runneruplist = qa_dict
            break
    # If the list is greater than 1, and the second item has a score greater than .5, send the first and second items
    if len(dictlist) > 1:
        return jsonify({'Best': dictlist[0], 'Second': dictlist[1]})
    # If the list is 1, send the first item
    elif len(dictlist) == 1:
        return jsonify({'Best': dictlist[0]})



# Define the main method
def main(prompt):

    with tqdm(total=4, desc="Finding Keywords") as pbar:
        # Get the user input from get_promt()
        #keywords = get_keywords(prompt)
        #print(f"Searching for '{keywords}'...")
        pbar.update(1)
        pbar.set_description("Obtaining Stack Questions")

        # Get the questions and answers
        search_data, highest_score_answers = get_questions(prompt)
        pbar.update(1)
        pbar.set_description("Scoring Stack Responses")

        # Get the relevant response
        dictlist = get_relevant_response(prompt, search_data, highest_score_answers)
        pbar.update(1)
        pbar.set_description("Saving and Formatting Responses")
    
        print("prompt"+prompt)

        # Create an excel file
        create_excel(dictlist, prompt)
        pbar.update(1)
    # Print a message to the user
    print("Excel file created successfully.")

    # Return the questions and answers
    return get_best_questions(dictlist)


# Create the API endpoint
@app.route('/stack_responses', methods=['POST'])
def stack_responses():
    # Get the request data
    request_data = request.get_json()

    # Extract the prompt from the request data
    prompt = request_data['prompt']
    
    # Call the main method
    dictlist = main(prompt)

    # Return the results as JSON
    return dictlist

if __name__ == '__main__':
    # Run the Flask app in a separate thread
    target=app.run(debug=True,host='localhost', port=5050)